from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
from openpyxl import load_workbook
from PIL import Image
import os

# Load the Excel sheet
wb = load_workbook('data.xlsx')
ws = wb.active

# Set up table style
table_style = TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, 0), 14),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ('GRID', (0, 0), (-1, -1), 1, colors.black),
])

# Set up the image size and position
photo_width = 100
photo_height = 100
photo_x = 50
photo_y = 200

# Get the current working directory
cwd = os.getcwd()


def max_row():
    global_max_row = ws.max_row
    if global_max_row is None:
        return 0  # Set a default value when max_row() returns None
    for row in range(1, global_max_row+1):
        if str(ws["A"+str(row)].value) == "None":
            return (row - 1)
    return global_max_row


# Iterate over the rows in the Excel sheet
for row in ws.iter_rows(min_row=2, max_row=max_row(), values_only=True):

    name = row[0]
    standard = row[1]
    division = row[2]
    phone = row[3]
    photo_filename = row[4]
    print(name, photo_filename, phone)

    # Create a new PDF for each student
    pdf = canvas.Canvas(f'ID Card\{name}.pdf', pagesize=letter)

    # Load the ID card template background
    template_path = 'template.jpg'  # Update with the path to your template background
    pdf.drawImage(template_path, 0, 0, width=letter[0], height=letter[1])

    # Draw the ID card layout
    pdf.setFont('Helvetica-Bold', 24)
    pdf.drawString(50, 700, f'{name}')
    pdf.drawString(50, 670, f'{standard}')
    pdf.drawString(50, 640, f'{division}')
    pdf.drawString(50, 610, f'{phone}')

    # Construct the photo file path
    photo_path = os.path.join(cwd, 'photos', photo_filename)
    print(photo_path)

    # Load and resize the photo
    # photo = Image.open(photo_path)
    # photo = photo.resize((photo_width, photo_height), Image.ANTIALIAS)
    photo = ImageReader(photo_path)

    # Add the photo to the PDF
    pdf.drawImage(
        ImageReader(photo),
        photo_x,
        photo_y,
        width=photo_width,
        height=photo_height,
        preserveAspectRatio=True,
    )

    # Save the PDF file
    pdf.save()
